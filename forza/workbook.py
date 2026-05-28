"""Builds the polished, sortable Excel tier list from the data model.

Design goals (the "1000x better than a flat PDF" brief):
  * Sortable + filterable master table (Excel AutoFilter, frozen header).
  * Tier-band colour coding by class, confidence-coloured cells, discipline banding.
  * A scannable discipline x class quick-reference matrix.
  * A real Tuning sheet: per-discipline baselines, a tune-codes table, per-car
    specs, and a code column that is honest about provenance.
  * A Sources sheet documenting trust and methodology.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import seed
from .models import (
    CAR_CLASSES,
    CLASS_COLORS,
    CLASS_PI_RANGE,
    CONFIDENCE_COLORS,
    DISCIPLINES,
    Car,
    class_rank,
)

# --- Palette -----------------------------------------------------------------
INK = "1F2A37"
HEADER_FILL = PatternFill("solid", fgColor=INK)
TITLE_FILL = PatternFill("solid", fgColor="111827")
BAND_FILL = PatternFill("solid", fgColor="F3F4F6")
ZEBRA_FILL = PatternFill("solid", fgColor="F5F7FA")  # alternating discipline band

# Per-sheet tab colours for a polished workbook.
TAB_COLORS = {
    "Overview": "111827", "Tier List": "2563EB", "Quick Reference": "059669",
    "Tuning": "D97706", "Sources": "6B7280",
}
WHITE = Font(color="FFFFFF", bold=True)
WHITE_BIG = Font(color="FFFFFF", bold=True, size=20)
WHITE_SUB = Font(color="D1D5DB", size=11)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _readable_text(hex_bg: str) -> str:
    """Pick black/white text for contrast against a coloured cell."""
    r, g, b = int(hex_bg[0:2], 16), int(hex_bg[2:4], 16), int(hex_bg[4:6], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b)
    return "000000" if luminance > 150 else "FFFFFF"


def _style_header(ws: Worksheet, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, text in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=text)
        c.fill = HEADER_FILL
        c.font = WHITE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _set_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _sorted_cars(cars: list[Car]) -> list[Car]:
    disc_order = {d: i for i, d in enumerate(DISCIPLINES)}
    return sorted(
        cars,
        key=lambda c: (
            disc_order.get(c.discipline, 99),
            -class_rank(c.car_class),  # R (highest) first
            c.rank,
        ),
    )


# --- Sheets ------------------------------------------------------------------

def _overview_sheet(wb: Workbook, meta: dict, cars: list[Car]) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS["Overview"]
    _set_widths(ws, {1: 22, 2: 90})

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"{meta['game']} — Best Cars by Tier & Discipline"
    t.fill = TITLE_FILL
    t.font = WHITE_BIG
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:B2")
    s = ws["A2"]
    s.value = meta.get("snapshot", "")
    s.fill = TITLE_FILL
    s.font = WHITE_SUB
    ws.row_dimensions[2].height = 20

    n_disc = len({c.discipline for c in cars})
    coverage = f"{len(cars)} cars across {n_disc} disciplines, classes D through R."
    rows = [
        ("Edition", meta.get("edition_note", "")),
        ("Coverage", coverage),
        ("PI bands", meta.get("pi_bands", "")),
        ("Last updated", meta.get("last_updated", "")),
        ("Read this first", meta.get("disclaimer", "")),
    ]
    r = 4
    for label, val in rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = BOLD
        lc.alignment = WRAP
        vc = ws.cell(row=r, column=2, value=val)
        vc.alignment = WRAP
        ws.row_dimensions[r].height = 30 if len(str(val)) > 80 else 18
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Confidence key").font = Font(bold=True, size=13)
    r += 1
    for level, desc in seed.CONFIDENCE_KEY.items():
        c = ws.cell(row=r, column=1, value=level)
        c.fill = PatternFill("solid", fgColor=CONFIDENCE_COLORS[level])
        c.font = Font(bold=True, color=_readable_text(CONFIDENCE_COLORS[level]))
        c.alignment = CENTER
        c.border = BORDER
        ws.cell(row=r, column=2, value=desc).alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Class tiers (low -> high)").font = Font(bold=True, size=13)
    r += 1
    for i, cls in enumerate(CAR_CLASSES):
        c = ws.cell(row=r, column=1 + i, value=cls)
        c.fill = PatternFill("solid", fgColor=CLASS_COLORS[cls])
        c.font = Font(bold=True, color=_readable_text(CLASS_COLORS[cls]))
        c.alignment = CENTER
        c.border = BORDER

    r += 2
    ws.cell(row=r, column=1, value="Sheets").font = Font(bold=True, size=13)
    r += 1
    guide = [
        ("Tier List", "Every ranked car with its in-game PI number. Use the filter arrows to sort by class, PI, price, confidence, or discipline."),
        ("Quick Reference", "Discipline x class grid — the top pick in each cell, colour coded."),
        ("Tuning", "Per-discipline baseline setups, the FH6 Mechanical Balance stat, per-car tune notes, and where to get real tune codes (marked *)."),
        ("Sources", "What was cross-referenced, and which low-quality sites to distrust."),
    ]
    for name, desc in guide:
        ws.cell(row=r, column=1, value=name).font = BOLD
        ws.cell(row=r, column=2, value=desc).alignment = WRAP
        r += 1

    r += 1
    note = ws.cell(row=r, column=1, value="* on a tune code")
    note.font = Font(bold=True, color="B45309")
    n2 = ws.cell(row=r, column=2, value="Community-sourced, not verified — confirm it in-game before trusting it. See the Tuning sheet.")
    n2.alignment = WRAP


def _tier_list_sheet(wb: Workbook, cars: list[Car]) -> None:
    ws = wb.create_sheet("Tier List")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS["Tier List"]
    headers = ["Discipline", "Class", "#", "Car", "Year", "PI band", "Price (CR)",
               "Acquisition", "Confidence", "Strengths", "Weaknesses", "Sources"]
    _style_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    disc_order = {d: i for i, d in enumerate(DISCIPLINES)}
    row = 2
    for car in _sorted_cars(cars):
        # Alternate a faint band per discipline group so disciplines read as blocks.
        banded = disc_order.get(car.discipline, 0) % 2 == 1
        pi_display = str(car.pi_value) if car.pi_value else CLASS_PI_RANGE.get(car.car_class, "")
        vals = [
            car.discipline, car.car_class, car.rank, car.display_name(),
            car.year or "", pi_display, car.price_cr if car.price_cr else "",
            car.acquisition, car.confidence, car.strengths, car.weaknesses,
            car.sources,
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.border = BORDER
            c.alignment = WRAP if i in (4, 8, 10, 11, 12) else CENTER
            if banded and i not in (2, 9):  # skip the class/confidence colour cells
                c.fill = ZEBRA_FILL
        # Colour the Class cell
        cls_cell = ws.cell(row=row, column=2)
        bg = CLASS_COLORS.get(car.car_class, "FFFFFF")
        cls_cell.fill = PatternFill("solid", fgColor=bg)
        cls_cell.font = Font(bold=True, color=_readable_text(bg))
        # Colour the Confidence cell
        conf_cell = ws.cell(row=row, column=9)
        cbg = CONFIDENCE_COLORS.get(car.confidence, "FFFFFF")
        conf_cell.fill = PatternFill("solid", fgColor=cbg)
        conf_cell.font = Font(bold=True, color=_readable_text(cbg))
        # Price formatting
        if car.price_cr:
            ws.cell(row=row, column=7).number_format = "#,##0"
        ws.row_dimensions[row].height = 42
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    _set_widths(ws, {1: 17, 2: 7, 3: 4, 4: 34, 5: 7, 6: 9, 7: 12,
                     8: 30, 9: 13, 10: 48, 11: 40, 12: 24})


def _quick_reference_sheet(wb: Workbook, cars: list[Car]) -> None:
    ws = wb.create_sheet("Quick Reference")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS["Quick Reference"]
    classes_desc = list(reversed(CAR_CLASSES))  # R first
    _style_header(ws, 1, ["Discipline / Class"] + classes_desc)
    ws.freeze_panes = "B2"

    # Index best (rank 1, else lowest rank) car per discipline+class.
    best: dict[tuple[str, str], Car] = {}
    for c in cars:
        key = (c.discipline, c.car_class)
        if key not in best or c.rank < best[key].rank:
            best[key] = c

    row = 2
    for disc in DISCIPLINES:
        dc = ws.cell(row=row, column=1, value=disc)
        dc.font = BOLD
        dc.fill = BAND_FILL
        dc.alignment = WRAP
        dc.border = BORDER
        for i, cls in enumerate(classes_desc, start=2):
            car = best.get((disc, cls))
            cell = ws.cell(row=row, column=i, value=car.display_name() if car else "—")
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = BORDER
            if car:
                bg = CLASS_COLORS[cls]
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(color=_readable_text(bg), size=10, bold=True)
        ws.row_dimensions[row].height = 46
        row += 1

    _set_widths(ws, {1: 20, **{i: 23 for i in range(2, 2 + len(classes_desc))}})


def _tuning_sheet(wb: Workbook, cars: list[Car]) -> None:
    ws = wb.create_sheet("Tuning")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS["Tuning"]
    _set_widths(ws, {1: 22, 2: 30, 3: 26, 4: 30, 5: 30})

    ws.merge_cells("A1:E1")
    h = ws["A1"]
    h.value = "Tuning — reliable specs, not fake codes"
    h.fill = TITLE_FILL
    h.font = WHITE_BIG
    ws.row_dimensions[1].height = 30

    # Baseline matrix
    row = 3
    _style_header(ws, row, ["Setting", "Road / Grip", "Rally / Off-road", "Drift", "Drag"])
    row += 1
    for b in seed.TUNE_BASELINES:
        vals = [b["setting"], b["road"], b["rally"], b["drift"], b["drag"]]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.alignment = WRAP
            c.border = BORDER
            if i == 1:
                c.font = BOLD
                c.fill = BAND_FILL
        ws.row_dimensions[row].height = 38
        row += 1

    row += 1
    for label, text in [("Mechanical Balance", seed.MECH_BALANCE),
                        ("Recommended tune order", seed.TUNE_ORDER)]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        m = ws.cell(row=row, column=2, value=text)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        m.alignment = WRAP
        ws.row_dimensions[row].height = 46 if len(text) > 90 else 20
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Universal tips").font = Font(bold=True, size=13)
    row += 1
    for tip in seed.UNIVERSAL_TIPS:
        t = ws.cell(row=row, column=1, value="• " + tip)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        t.alignment = WRAP
        ws.row_dimensions[row].height = 28
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Where to get tune codes").font = Font(bold=True, size=13)
    row += 1
    for line in seed.HOW_TO_GET_CODES:
        t = ws.cell(row=row, column=1, value=line)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        t.alignment = WRAP
        ws.row_dimensions[row].height = 22
        row += 1
    for label, url in seed.CODE_LINKS:
        c = ws.cell(row=row, column=1, value=label)
        c.hyperlink = url
        c.font = Font(color="0563C1", underline="single")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 20
        row += 1

    # The * legend
    row += 1
    leg = ws.cell(row=row, column=1, value=seed.CODE_LEGEND)
    leg.font = Font(italic=True, color="B45309")
    leg.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 60

    # Real share codes table (community-sourced)
    if seed.TUNE_CODES:
        row += 2
        ws.cell(row=row, column=1, value="Tune share codes (community-sourced *)").font = Font(bold=True, size=13)
        row += 1
        _style_header(ws, row, ["Car", "PI", "Share code *", "Best for", "Source"])
        row += 1
        for code in seed.TUNE_CODES:
            vals = [code.get("car", ""), code.get("pi", ""), code.get("code", ""),
                    code.get("best_for", ""), code.get("source", "")]
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=i, value=v)
                c.alignment = WRAP if i in (1, 4) else CENTER
                c.border = BORDER
                if i == 3:
                    c.font = Font(bold=True, name="Consolas")
            ws.row_dimensions[row].height = 24
            row += 1

    # Per-car tune table
    row += 2
    ws.cell(row=row, column=1, value="Per-car tune notes & code source").font = Font(bold=True, size=13)
    row += 1
    _style_header(ws, row, ["Car", "Discipline", "Class", "Tune notes", "Tune code source *"])
    header_row = row
    row += 1
    for car in _sorted_cars(cars):
        code_text = car.share_code if car.share_code else (car.code_status or seed.DEFAULT_CODE_STATUS)
        vals = [car.display_name(), car.discipline, car.car_class,
                car.tune_summary, code_text]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.alignment = WRAP
            c.border = BORDER
            if i == 3:
                bg = CLASS_COLORS.get(car.car_class, "FFFFFF")
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(bold=True, color=_readable_text(bg))
                c.alignment = CENTER
        ws.row_dimensions[row].height = 44
        row += 1
    ws.auto_filter.ref = f"A{header_row}:E{row - 1}"


def _sources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS["Sources"]
    _set_widths(ws, {1: 50, 2: 60})

    ws.merge_cells("A1:B1")
    h = ws["A1"]
    h.value = "Sources & methodology"
    h.fill = TITLE_FILL
    h.font = WHITE_BIG
    ws.row_dimensions[1].height = 30

    row = 3
    ws.cell(row=row, column=1, value="Cross-referenced (trusted)").font = Font(bold=True, size=13)
    row += 1
    for src in seed.SOURCES_TRUSTED:
        c = ws.cell(row=row, column=1, value="• " + src)
        c.alignment = WRAP
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Treated with caution (SEO farms / auto-generated)").font = Font(bold=True, size=13, color="B91C1C")
    row += 1
    for src in seed.SOURCES_CAUTION:
        c = ws.cell(row=row, column=1, value="• " + src)
        c.font = Font(color="B91C1C")
        row += 1
    ws.cell(row=row, column=1, value="(Cross-checked only — never used as a sole basis. Fabricated 'telemetry' disregarded.)").alignment = WRAP
    row += 2

    ws.cell(row=row, column=1, value="Confidence key").font = Font(bold=True, size=13)
    row += 1
    for level, desc in seed.CONFIDENCE_KEY.items():
        c = ws.cell(row=row, column=1, value=level)
        c.fill = PatternFill("solid", fgColor=CONFIDENCE_COLORS[level])
        c.font = Font(bold=True, color=_readable_text(CONFIDENCE_COLORS[level]))
        c.alignment = CENTER
        c.border = BORDER
        d = ws.cell(row=row, column=2, value=desc)
        d.alignment = WRAP
        row += 1


def build_workbook(cars: list[Car], meta: dict, out_path: str) -> str:
    wb = Workbook()
    _overview_sheet(wb, meta, cars)
    _tier_list_sheet(wb, cars)
    _quick_reference_sheet(wb, cars)
    _tuning_sheet(wb, cars)
    _sources_sheet(wb)
    wb.save(out_path)
    return out_path
